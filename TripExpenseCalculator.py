from ollama import chat
import pandas as pd
from datetime import datetime, timedelta
from dateutil import parser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def analyze_trip_expenses(start_date, end_date, location, accommodation, csv_files):
    """
    Analyzes trip expenses from financial statements using a local AI model.

    Args:
        start_date (str): Start date of the trip (YYYY-MM-DD).
        end_date (str): End date of the trip (YYYY-MM-DD).
        location (str): Location of the trip.
        accommodation (str): Type of accommodation (hotel, resort, cruise, etc.).
        csv_files (list): List of paths to CSV files containing financial statements.

    Returns:
        dict: A dictionary containing categorized expenses and total trip expenditure.
    """

    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.strptime(end_date, "%Y-%m-%d")
    advance_booking_months = 3  # Check for advance bookings up to 1 month prior

    all_expenses = []

    for csv_file in csv_files:
        print(f"Processing file: {csv_file}")
        print("-" * 20)
        print("")

        df = pd.read_csv(csv_file)

        columnListFromCsv = df.columns.tolist()
        
        # Identify relevant date, description and amount columns
        date_column = get_column_name(columnListFromCsv, "transaction date")
        description_column = get_column_name(columnListFromCsv, "name of the merchant")
        amount_column = get_column_name(columnListFromCsv, "transaction amount")

        print(f"Relevant columns from CSV detected. Using columns {date_column} for Date, {description_column} for Description, {amount_column} for Amount")

        # Assuming CSV has columns: 'date', 'description', 'amount'
        df['date'] = pd.to_datetime(df[date_column])

        for index, row in df.iterrows():
            transaction_date = row[date_column]
            description = row[description_column]
            amount = abs(row[amount_column]) 
            transaction_date_parsed = parser.parse(transaction_date) 
            category = "Other"

            # Filter transactions within the trip dates or for advance bookings
            if transaction_date_parsed >= start_date and transaction_date_parsed <= end_date:
                # Categorize expenses
                category = categorize_expense(description, location, amount)

            elif transaction_date_parsed >= (start_date - timedelta(days=advance_booking_months * 30)) and transaction_date_parsed <= (start_date - timedelta(days=1)):
                # Categorize expenses for advance bookings. We will look for specific types here (airfare, accommodation)
                category = categorize_expense(description, location, amount)

                if(category not in ["Airfare", "Accommodation"]):
                    #log_transaction(transaction_date_parsed, description, amount, category, status="Ignored", sub_status="Older irrelevant transaction")
                    continue  

            else:
                # Ignore transactions outside the trip date range and advance booking window
                continue
    
            # ignore subscriptions and recurring payments
            if (category not in ["Subscriptions", "Recurring Payments"]):
                log_transaction(transaction_date_parsed, description, amount, category, status="Accepted")
                all_expenses.append({'date': transaction_date, 'description': description, 'amount': amount, 'category': category})
            #else:
                #log_transaction(transaction_date_parsed, description, amount, category, status="Ignored", sub_status="Subscription/Recurring Payment")

        print("")

    # Print all expenses by category
    print_expenses_by_category(all_expenses)

    # Group expenses by category and calculate sums
    categorized_expenses = {}
    for expense in all_expenses:
        category = expense['category']
        if category not in categorized_expenses:
            categorized_expenses[category] = 0
        categorized_expenses[category] += expense['amount']

    # Calculate total trip expenditure
    total_expenditure = sum(categorized_expenses.values())

    write_expenses_to_excel(all_expenses, categorized_expenses, total_expenditure, "trip_expenses.xlsx")

    return categorized_expenses, total_expenditure

def categorize_expense(description, location, amount):
    """
    Categorizes an expense based on its description and location.
    Uses Ollama to categorize expenses.
    """
    messages = [
        {
            'role': 'user',
            'content': f"Categorize the following expense description: '{description}' for a trip to {location}, for an amount of {amount}. Possible categories are: Accommodation, Transport, Meals, Tours, Groceries, Airfare, Subscriptions, Recurring Payments, Other. Pay special attention to the description, if you find full or partial names of airline companies and expenses are on the higher side (say double digits to hundreds of dollars), then it is likely Airfare. Print only the category and nothing else.",
        },
    ]
    response = chat('gemma3:12b', messages=messages)
    category = response['message']['content'].strip()
    return category

def get_column_name(columnList, requiredColumnDescription):
    """
    Given a list of column names and a required column description,
    returns the best matching column name.
    """

    concatenatedColumnList = ",".join(columnList)

    messages = [
        {
            'role': 'user',
            'content': f"Among these columns in a credit card/banking statement, which one refers to the one with the '{requiredColumnDescription}'? {concatenatedColumnList}. Just list the name of the column.",
        },
    ]

    response = chat('gemma3:12b', messages=messages)
    category = response['message']['content'].strip()
    return category

def log_transaction(transaction_date, description, amount, category, status, sub_status=None):
    """
    Logs a bank account/credit card statement transaction.

    Args:
        transaction_date (datetime): The date of the transaction.
        description (str): The description of the transaction.
        amount (float): The amount of the transaction.
        category (str): The category of the transaction.
        status (str): The status of the transaction ('Accepted' or 'Ignored').
        sub_status (str, optional): A sub-status if the transaction is ignored. Defaults to None.
    """

    log_entry = f"{transaction_date.strftime('%Y-%m-%d')}|{description}|{amount}|{category}|{status}"
    if sub_status:
        log_entry += f"|{sub_status}"

    print(log_entry)

def log_finalized_transaction(transaction_date, description, amount):
    """
    Logs a bank account/credit card statement transaction.

    Args:
        transaction_date (datetime): The date of the transaction.
        description (str): The description of the transaction.
        amount (float): The amount of the transaction.
    """

    log_entry = f"{transaction_date.strftime('%Y-%m-%d')}|{description}|{amount}"
    print(log_entry)

def print_expenses_by_category(all_expenses):
    """
    Prints out all expenses in the all_expenses array, category by category, sorted by date.

    Args:
        all_expenses (list): A list of expenses, each assigned a category.
    """

    categories = sorted(list(set([expense['category'] for expense in all_expenses])))

    print("-" * 20)
    print("-" * 20)
    for category in categories:
        expenses_in_category = sorted([expense for expense in all_expenses if expense['category'] == category], key=lambda x: x['date'])
        print(f"Category: {category}")
        for expense in expenses_in_category:
            log_finalized_transaction(parser.parse(expense['date']), expense['description'], expense['amount'])
        print("-" * 20)


def write_expenses_to_excel(all_expenses, categorized_expenses, total_expenditure, output_file):
    """
    Writes categorized expenses and total expenditure to an Excel file.
    
    Args:
        all_expenses (list): A list of all expense dictionaries with date, description, amount, and category.
        categorized_expenses (dict): A dictionary of category totals.
        total_expenditure (float): The total trip expenditure.
        output_file (str): The path to the output Excel file.
    """
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_font = Font(bold=True)
    total_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    total_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create a sheet for each category
    categories = sorted(list(set([expense['category'] for expense in all_expenses])))
    
    for category in categories:
        ws = wb.create_sheet(title=category[:31])  # Excel sheet name limit is 31 characters
        
        # Add header row
        headers = ['Date', 'Description', 'Amount']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add transactions for this category
        expenses_in_category = sorted([expense for expense in all_expenses if expense['category'] == category], key=lambda x: x['date'])
        for row, expense in enumerate(expenses_in_category, start=2):
            ws.cell(row=row, column=1, value=expense['date']).border = border
            ws.cell(row=row, column=2, value=expense['description']).border = border
            ws.cell(row=row, column=3, value=expense['amount']).border = border
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
        
        # Add category total
        total_row = len(expenses_in_category) + 2
        ws.cell(row=total_row, column=1, value='Total').font = total_font
        ws.cell(row=total_row, column=2, value=category).font = total_font
        total_cell = ws.cell(row=total_row, column=3, value=categorized_expenses[category])
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.number_format = '$#,##0.00'
        total_cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
    
    # Create summary sheet
    summary_ws = wb.create_sheet(title='Summary', index=0)
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 15
    
    # Add title
    summary_ws.cell(row=1, column=1, value='Trip Expense Summary')
    summary_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    # Add header row for summary table
    summary_ws.cell(row=3, column=1, value='Category')
    summary_ws.cell(row=3, column=2, value='Total Amount')
    for col in [1, 2]:
        cell = summary_ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add category totals
    for row, (category, total) in enumerate(sorted(categorized_expenses.items()), start=4):
        summary_ws.cell(row=row, column=1, value=category).border = border
        amount_cell = summary_ws.cell(row=row, column=2, value=total)
        amount_cell.border = border
        amount_cell.number_format = '$#,##0.00'
    
    # Add grand total
    grand_total_row = len(categorized_expenses) + 4
    summary_ws.cell(row=grand_total_row, column=1, value='Grand Total').font = total_font
    grand_total_cell = summary_ws.cell(row=grand_total_row, column=2, value=total_expenditure)
    grand_total_cell.font = total_font
    grand_total_cell.fill = total_fill
    grand_total_cell.number_format = '$#,##0.00'
    grand_total_cell.border = border
    
    # Save the workbook
    wb.save(output_file)
    print(f"Expenses written to {output_file}")


# Example invocation
if __name__ == "__main__":
    start_date = "2025-12-16"
    end_date = "2025-12-23"
    location = "Cozumel, Mexico"
    accommodation = "Resort and Hotel"
    csv_files = ["C:\\mine\\TripExpenseData\\FirstTechExportedTransactions.csv", "C:\\mine\\TripExpenseData\\Chase Credit Card.CSV", "C:\\mine\\TripExpenseData\\Fidelity Credit Card.CSV"]

    categorized_expenses, total_expenditure = analyze_trip_expenses(start_date, end_date, location, accommodation, csv_files)

    print("Categorized Expenses:")
    print(categorized_expenses)
    print("\nTotal Expenditure:", total_expenditure)

    