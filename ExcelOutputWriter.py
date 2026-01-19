from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from OutputWriter import OutputWriter

class ExcelOutputWriter(OutputWriter):
    """
    Concrete implementation of OutputWriter for writing to an Excel file.
    """
    def __init__(self, excel_file_name):
        """
        Initializes the ExcelOutputWriter with the output file name.

        Args:
            excel_file_name (str): The path to the output Excel file.
        """
        self.excel_file_name = excel_file_name
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

        # Define styles (moved to write_expenses for consistency)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.total_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.total_font = Font(bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.alignment = Alignment(horizontal='center', vertical='center')


    def write_expenses(self, all_expenses, categorized_expenses, total_expenditure):
        """
        Writes categorized expenses and total expenditure to an Excel file.

        Args:
            all_expenses (list): A list of all expense dictionaries with date, description, amount, and category.
            categorized_expenses (dict): A dictionary of category totals.
            total_expenditure (float): The total trip expenditure.
        """

        # Create a sheet for each category
        categories = sorted(list(set([expense['category'] for expense in all_expenses])))

        for category in categories:
            self.print_expenses_by_category(category, all_expenses, categorized_expenses)

        # Create summary sheet
        summary_ws = self.wb.create_sheet(title='Summary', index=0)
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 15

        # Add title
        summary_ws.cell(row=1, column=1, value='Trip Expense Summary')
        summary_ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Add header row for summary table
        summary_ws.cell(row=3, column=1, value='Category')
        summary_ws.cell(row=3, column=2, value='Total Amount')
        for col in [1, 2]:
            cell = summary_ws.cell(row=3, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.alignment

        # Add category totals
        for row, (category, total) in enumerate(sorted(categorized_expenses.items()), start=4):
            summary_ws.cell(row=row, column=1, value=category).border = self.border
            amount_cell = summary_ws.cell(row=row, column=2, value=total)
            amount_cell.border = self.border
            amount_cell.number_format = '$#,##0.00'

        # Add grand total
        grand_total_row = len(categorized_expenses) + 4
        summary_ws.cell(row=grand_total_row, column=1, value='Grand Total').font = self.total_font
        grand_total_cell = summary_ws.cell(row=grand_total_row, column=2, value=total_expenditure)
        grand_total_cell.font = self.total_font
        grand_total_cell.fill = self.total_fill
        grand_total_cell.number_format = '$#,##0.00'
        grand_total_cell.border = self.border

        self.wb.save(self.excel_file_name)
        print(f"Expenses written to {self.excel_file_name}")

    def print_expenses_by_category(self, category, all_expenses, categorized_expenses):
        """Outputs expenses in a specific category to excel."""

        ws = self.wb.create_sheet(title=category[:31])  # Excel sheet name limit is 31 characters

        # Add header row
        headers = ['Date', 'Description', 'Amount']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.alignment

        # Add transactions for this category
        expenses_in_category = sorted([expense for expense in all_expenses if expense['category'] == category], key=lambda x: x['date'])
        for row, expense in enumerate(expenses_in_category, start=2):
            ws.cell(row=row, column=1, value=expense['date']).border = self.border
            ws.cell(row=row, column=2, value=expense['description']).border = self.border
            ws.cell(row=row, column=3, value=expense['amount']).border = self.border
            ws.cell(row=row, column=3).number_format = '$#,##0.00'

        # Add category total
        total_row = len(expenses_in_category) + 2
        ws.cell(row=total_row, column=1, value='Total').font = self.total_font
        ws.cell(row=total_row, column=2, value=category).font = self.total_font
        total_cell = ws.cell(row=total_row, column=3, value=categorized_expenses[category])
        total_cell.font = self.total_font
        total_cell.fill = self.total_fill
        total_cell.number_format = '$#,##0.00'
        total_cell.border = self.border

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12